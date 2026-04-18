"""
Export helpers — turn a list of data tables into Markdown text or an XLSX file.

Each caller provides a uniform `tables` structure:

    tables = [
        {
            "name": "Pageviews",
            "columns": ["Page", "Views", "Users"],
            "rows": [
                ["/home", 120, 80],
                ["/about", 45, 30],
            ],
        },
        ...
    ]

Optionally a table may include a "summary" dict (rendered in Markdown as a small
key-value block above the table; included in XLSX as header rows).
"""

from io import BytesIO
from typing import List, Dict, Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def to_markdown(title: str, tables: List[Dict[str, Any]], subtitle: str = None) -> str:
    lines: List[str] = [f"# {title}"]
    if subtitle:
        lines.append(f"_{subtitle}_")
    lines.append("")

    for table in tables:
        lines.append(f"## {table['name']}")
        lines.append("")

        summary = table.get("summary")
        if summary:
            for key, value in summary.items():
                lines.append(f"- **{key}:** {value}")
            lines.append("")

        columns = table.get("columns") or []
        rows = table.get("rows") or []

        if not columns:
            continue

        lines.append("| " + " | ".join(str(c) for c in columns) + " |")
        lines.append("| " + " | ".join("---" for _ in columns) + " |")

        if not rows:
            lines.append("| " + " | ".join("—" for _ in columns) + " |")
        else:
            for row in rows:
                cells = [_md_cell(c) for c in row]
                lines.append("| " + " | ".join(cells) + " |")

        lines.append("")

    return "\n".join(lines).rstrip() + "\n"


def _md_cell(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    return text.replace("|", "\\|").replace("\n", " ")


def to_xlsx(tables: List[Dict[str, Any]]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="F0F0F0")

    used_names: Dict[str, int] = {}

    for table in tables:
        sheet_name = _safe_sheet_name(table["name"], used_names)
        ws = wb.create_sheet(title=sheet_name)

        current_row = 1

        summary = table.get("summary")
        if summary:
            for key, value in summary.items():
                ws.cell(row=current_row, column=1, value=str(key)).font = header_font
                ws.cell(row=current_row, column=2, value=_xlsx_value(value))
                current_row += 1
            current_row += 1

        columns = table.get("columns") or []
        rows = table.get("rows") or []

        if columns:
            for col_idx, col_name in enumerate(columns, start=1):
                cell = ws.cell(row=current_row, column=col_idx, value=str(col_name))
                cell.font = header_font
                cell.fill = header_fill
            header_row = current_row
            current_row += 1

            for row in rows:
                for col_idx, value in enumerate(row, start=1):
                    ws.cell(row=current_row, column=col_idx, value=_xlsx_value(value))
                current_row += 1

            for col_idx, col_name in enumerate(columns, start=1):
                max_len = len(str(col_name))
                for row in rows:
                    if col_idx - 1 < len(row):
                        cell_len = len(str(row[col_idx - 1] if row[col_idx - 1] is not None else ""))
                        if cell_len > max_len:
                            max_len = cell_len
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _safe_sheet_name(name: str, used: Dict[str, int]) -> str:
    cleaned = name[:31]
    for bad in ("/", "\\", "*", "?", ":", "[", "]"):
        cleaned = cleaned.replace(bad, "-")
    if not cleaned:
        cleaned = "Sheet"
    if cleaned in used:
        used[cleaned] += 1
        suffix = f" ({used[cleaned]})"
        cleaned = (cleaned[: 31 - len(suffix)]) + suffix
    else:
        used[cleaned] = 1
    return cleaned


def _xlsx_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (int, float, bool)):
        return value
    return str(value)
